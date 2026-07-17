"""Logs successful user registrations to logs/user_activity.xlsx (one row per new account)."""
import json
import os
import threading
from datetime import datetime

from django.conf import settings
from django.utils import timezone as dj_timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOG_DIR = os.path.join(settings.BASE_DIR, 'logs')
LOG_FILE = os.path.join(LOG_DIR, 'user_activity.xlsx')
# Rows that couldn't be saved (e.g. the workbook was open/locked in Excel at
# the time) land here and get merged into the sheet on the next successful save.
PENDING_FILE = os.path.join(LOG_DIR, '.pending_activity.jsonl')
SHEET_NAME = 'User Activity'

HEADERS = [
    'Date', 'Time', 'Full Name', 'Email', 'Role', 'Education Type',
    'Action', 'IP Address', 'Browser', 'Operating System', 'Device Type', 'Login Status',
]

# Excel isn't safe for concurrent writers, so serialize access within this process.
_lock = threading.Lock()


def _get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def _parse_user_agent(user_agent):
    ua = (user_agent or '').lower()

    if 'edg/' in ua:
        browser = 'Edge'
    elif 'opr/' in ua or 'opera' in ua:
        browser = 'Opera'
    elif 'chrome' in ua:
        browser = 'Chrome'
    elif 'firefox' in ua:
        browser = 'Firefox'
    elif 'safari' in ua:
        browser = 'Safari'
    else:
        browser = 'Unknown'

    if 'windows' in ua:
        os_name = 'Windows'
    elif 'android' in ua:
        os_name = 'Android'
    elif 'iphone' in ua or 'ipad' in ua or 'ios' in ua:
        os_name = 'iOS'
    elif 'mac os' in ua or 'macintosh' in ua:
        os_name = 'macOS'
    elif 'linux' in ua:
        os_name = 'Linux'
    else:
        os_name = 'Unknown'

    if 'ipad' in ua or 'tablet' in ua:
        device_type = 'Tablet'
    elif 'mobi' in ua or 'android' in ua or 'iphone' in ua:
        device_type = 'Mobile'
    else:
        device_type = 'Desktop'

    return browser, os_name, device_type


def _write_header(sheet):
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')

    for col, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    sheet.freeze_panes = 'A2'


def _ensure_workbook():
    """Create the logs dir / workbook / sheet / header row if any are missing."""
    os.makedirs(LOG_DIR, exist_ok=True)

    if os.path.exists(LOG_FILE):
        workbook = load_workbook(LOG_FILE)
        if SHEET_NAME not in workbook.sheetnames:
            workbook.create_sheet(SHEET_NAME)
        sheet = workbook[SHEET_NAME]
        if sheet.cell(row=1, column=1).value != HEADERS[0]:
            _write_header(sheet)
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    _write_header(sheet)
    return workbook, sheet


def _style_row(sheet, row_index):
    center = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') if row_index % 2 == 0 else None

    for col in range(1, len(HEADERS) + 1):
        cell = sheet.cell(row=row_index, column=col)
        cell.alignment = center
        if fill:
            cell.fill = fill


def _queue_pending_rows(rows):
    os.makedirs(LOG_DIR, exist_ok=True)
    with open(PENDING_FILE, 'a', encoding='utf-8') as f:
        for row in rows:
            f.write(json.dumps(row) + '\n')


def _pop_pending_rows():
    if not os.path.exists(PENDING_FILE):
        return []
    with open(PENDING_FILE, 'r', encoding='utf-8') as f:
        rows = [json.loads(line) for line in f if line.strip()]
    os.remove(PENDING_FILE)
    return rows


def _peek_pending_rows():
    if not os.path.exists(PENDING_FILE):
        return []
    with open(PENDING_FILE, 'r', encoding='utf-8') as f:
        return [json.loads(line) for line in f if line.strip()]


def pending_count():
    if not os.path.exists(PENDING_FILE):
        return 0
    with open(PENDING_FILE, encoding='utf-8') as f:
        return sum(1 for line in f if line.strip())


def _autofit_columns(sheet):
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_length = max(
            (len(str(sheet.cell(row=r, column=col).value or '')) for r in range(1, sheet.max_row + 1)),
            default=len(HEADERS[col - 1]),
        )
        sheet.column_dimensions[letter].width = max(12, min(max_length + 2, 40))


def log_user_registration(user, request):
    """
    Append a row recording a successful registration to logs/user_activity.xlsx.
    One row per new account - logins are not logged.

    Never raises - failures are printed so they can't interrupt the registration
    flow that triggered them.
    """
    try:
        profile = getattr(user, 'profile', None)
        role = profile.get_role_display() if profile else ''
        education_type = profile.get_education_type_display() if profile and profile.education_type else ''

        ip_address = _get_client_ip(request)
        browser, os_name, device_type = _parse_user_agent(request.META.get('HTTP_USER_AGENT', ''))

        now = datetime.now()
        row = [
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            (user.first_name or user.username or '').strip(),
            user.email or '',
            role,
            education_type,
            'Register',
            ip_address,
            browser,
            os_name,
            device_type,
            'Success',
        ]

        with _lock:
            _write_rows_or_queue([row])

    except Exception as exc:
        print(f"[activity_logger] Failed to log user registration: {exc}")


def _write_rows_or_queue(new_rows):
    """
    Merge any previously-queued rows plus new_rows into the workbook and save.
    If the save fails (e.g. the file is open/locked in Excel), nothing is lost -
    every row involved goes back on the pending queue for the next attempt.
    """
    pending = _pop_pending_rows()
    rows_to_write = pending + list(new_rows)

    try:
        workbook, sheet = _ensure_workbook()
        for row in rows_to_write:
            sheet.append(row)
            _style_row(sheet, sheet.max_row)
        _autofit_columns(sheet)
        workbook.save(LOG_FILE)
    except Exception as exc:
        _queue_pending_rows(rows_to_write)
        raise exc


def sync_pending_entries():
    """Retry writing any rows that were queued because the workbook was locked."""
    with _lock:
        count = pending_count()
        if not count:
            return 0
        try:
            _write_rows_or_queue([])
            return count
        except Exception:
            return 0


def backfill_existing_users():
    """
    One-time import of users who registered before activity logging existed,
    so the sheet has exactly one Register row per account, not just the ones
    captured going forward.

    Per-request details (IP/browser/OS/device) were never captured for these
    accounts, so those columns are filled with 'Unknown'. Safe to re-run -
    skips any email that already has a Register row.
    """
    from django.contrib.auth.models import User

    with _lock:
        workbook, sheet = _ensure_workbook()
        already_logged = {
            row[3] for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True)
            if row[3] and row[6] == 'Register'
        }
        already_logged.update(
            row[3] for row in _peek_pending_rows() if row[3] and row[6] == 'Register'
        )

        new_rows = []
        for user in User.objects.select_related('profile').order_by('date_joined'):
            if not user.email or user.email in already_logged:
                continue

            profile = getattr(user, 'profile', None)
            role = profile.get_role_display() if profile else ''
            education_type = profile.get_education_type_display() if profile and profile.education_type else ''

            joined = user.date_joined
            if dj_timezone.is_aware(joined):
                joined = dj_timezone.localtime(joined)

            new_rows.append([
                joined.strftime('%Y-%m-%d'),
                joined.strftime('%H:%M:%S'),
                (user.first_name or user.username or '').strip(),
                user.email,
                role,
                education_type,
                'Register',
                'Unknown',
                'Unknown',
                'Unknown',
                'Unknown',
                'Success',
            ])

        if new_rows:
            _write_rows_or_queue(new_rows)

        return len(new_rows)
