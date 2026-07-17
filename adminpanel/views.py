from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .decorators import admin_required
from .forms import AdminCreateForm
from django.contrib.auth.models import User
from courses.models import Achievement, ActivityLog, Course, CourseSkill, Enrollment, LearningSession, LearningStreak, Skill, StreakReward, StudentProgress
from django.contrib import messages
from accounts.models import Profile
from accounts.notifications import notify, notify_all_students, notify_course_students
from courses.models import Lesson
from courses.models import Quiz, Question
from courses.models import Certificate
from datetime import date, datetime, time as dt_time, timezone as dt_timezone
from django.db.models import Count, F
from django.http import FileResponse, JsonResponse
from django.utils import timezone
import math
import os

from utils.activity_logger import LOG_FILE, backfill_existing_users, pending_count, sync_pending_entries

def _add_months(d, delta):
    m = d.month - 1 + delta
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, 1)

def _donut_segments(counts, radius=54):
    """Build SVG stroke-dasharray/dashoffset segments for a donut chart."""
    circumference = 2 * math.pi * radius
    total = sum(counts) or 1
    segments = []
    offset = 0.0
    for count in counts:
        fraction = count / total
        length = fraction * circumference
        gap = max(circumference - length - 1.5, 0) if length else circumference
        segments.append({
            'pct': round(fraction * 100),
            'dasharray': f"{max(length - (1.5 if length else 0), 0):.2f} {gap:.2f}",
            'dashoffset': f"{-offset:.2f}",
            'visible': length > 0,
        })
        offset += length
    return segments, circumference

def _simple_bar_series(rows, label_key, value_key, max_bars=6):
    """Cap/scale a list of {label, value} rows for a horizontal bar chart."""
    rows = list(rows)[:max_bars]
    max_value = max([r[value_key] for r in rows] + [1])
    return [
        {'label': r[label_key], 'value': r[value_key], 'pct': round(r[value_key] / max_value * 100)}
        for r in rows
    ], max_value

def _nice_ticks(max_val, target_count=4):
    """Round max_val up to a clean axis max and return evenly-spaced clean tick values."""
    if max_val <= 0:
        return [0], 1
    raw_step = max_val / (target_count - 1)
    magnitude = 10 ** math.floor(math.log10(raw_step))
    residual = raw_step / magnitude
    if residual > 5:
        nice_step = 10 * magnitude
    elif residual > 2:
        nice_step = 5 * magnitude
    elif residual > 1:
        nice_step = 2 * magnitude
    else:
        nice_step = magnitude
    nice_step = max(nice_step, 1)
    nice_max = math.ceil(max_val / nice_step) * nice_step
    ticks = [int(i * nice_step) for i in range(int(nice_max / nice_step) + 1)]
    return ticks, nice_max

def _smooth_path(points):
    """Monotone cubic Hermite spline (Fritsch-Carlson), converted to bezier.

    Plain Catmull-Rom smoothing overshoots past flat/repeated values (e.g. a
    run of zero counts dips below the baseline, implying negative data).
    Monotone interpolation is clamped so the curve never leaves the range of
    its neighboring points — same technique as D3's curveMonotoneX.
    """
    n = len(points)
    if n < 2:
        return ''
    xs = [p['x'] for p in points]
    ys = [p['y'] for p in points]
    dx = [xs[i + 1] - xs[i] for i in range(n - 1)]
    dy = [ys[i + 1] - ys[i] for i in range(n - 1)]
    slopes = [dy[i] / dx[i] if dx[i] else 0.0 for i in range(n - 1)]

    tangents = [0.0] * n
    tangents[0] = slopes[0]
    tangents[-1] = slopes[-1]
    for i in range(1, n - 1):
        if slopes[i - 1] * slopes[i] <= 0:
            tangents[i] = 0.0
        else:
            tangents[i] = (slopes[i - 1] + slopes[i]) / 2

    for i in range(n - 1):
        if slopes[i] == 0:
            tangents[i] = 0.0
            tangents[i + 1] = 0.0
        else:
            a = tangents[i] / slopes[i]
            b = tangents[i + 1] / slopes[i]
            h = math.hypot(a, b)
            if h > 3:
                t = 3 / h
                tangents[i] = t * a * slopes[i]
                tangents[i + 1] = t * b * slopes[i]

    path = f"M {xs[0]},{ys[0]}"
    for i in range(n - 1):
        cp1x = round(xs[i] + dx[i] / 3, 1)
        cp1y = round(ys[i] + tangents[i] * dx[i] / 3, 1)
        cp2x = round(xs[i + 1] - dx[i] / 3, 1)
        cp2y = round(ys[i + 1] - tangents[i + 1] * dx[i] / 3, 1)
        path += f" C {cp1x},{cp1y} {cp2x},{cp2y} {xs[i + 1]},{ys[i + 1]}"
    return path

@admin_required
def admin_dashboard(request):
    total_users = User.objects.count()
    total_courses = Course.objects.count()

    role_counts = {
        'STUDENT': User.objects.filter(profile__role='STUDENT').count(),
        'INSTRUCTOR': User.objects.filter(profile__role='INSTRUCTOR').count(),
        'ADMIN': User.objects.filter(profile__role='ADMIN').count(),
    }

    active_courses = Course.objects.filter(is_active=True).count()
    inactive_courses = total_courses - active_courses

    this_month = timezone.now().date().replace(day=1)
    months = [_add_months(this_month, -i) for i in range(5, -1, -1)]
    # Bucketed in Python (not via TruncMonth/__date) because this MySQL
    # instance has no time zone tables loaded, which makes CONVERT_TZ()
    # return NULL and silently drops every row from date-based DB queries.
    range_start_utc = datetime.combine(months[0], dt_time.min, tzinfo=dt_timezone.utc)
    trend_counts = {}
    for enrolled_at in Enrollment.objects.filter(enrolled_at__gte=range_start_utc).values_list('enrolled_at', flat=True):
        local_month = timezone.localtime(enrolled_at).date().replace(day=1)
        trend_counts[local_month] = trend_counts.get(local_month, 0) + 1
    enrollment_trend = [
        {'label': m.strftime('%b'), 'count': trend_counts.get(m, 0)}
        for m in months
    ]

    chart_w, chart_h = 960, 320
    pad_l, pad_r, pad_t, pad_b = 44, 24, 24, 36
    plot_w = chart_w - pad_l - pad_r
    plot_h = chart_h - pad_t - pad_b
    raw_max = max(e['count'] for e in enrollment_trend)
    y_ticks, axis_max = _nice_ticks(raw_max)
    step = plot_w / (len(enrollment_trend) - 1) if len(enrollment_trend) > 1 else 0
    baseline_y = pad_t + plot_h

    enrollment_points = []
    for i, e in enumerate(enrollment_trend):
        x = round(pad_l + i * step, 1)
        y = round(pad_t + (1 - e['count'] / axis_max) * plot_h, 1)
        enrollment_points.append({'x': x, 'y': y, 'count': e['count'], 'label': e['label']})

    line_points_str = _smooth_path(enrollment_points)
    area_points_str = (
        line_points_str +
        f" L {enrollment_points[-1]['x']},{baseline_y} L {enrollment_points[0]['x']},{baseline_y} Z"
    ) if enrollment_points else ''

    enrollment_gridlines = [
        {
            'y': round(pad_t + (1 - t / axis_max) * plot_h, 1),
            'label': t,
        }
        for t in y_ticks
    ]
    enrollment_last_point = enrollment_points[-1] if enrollment_points else None

    # Role distribution donut
    role_segments, role_circumference = _donut_segments([
        role_counts['STUDENT'], role_counts['INSTRUCTOR'], role_counts['ADMIN'],
    ])

    # Certificate status donut
    total_certificates = Certificate.objects.count()
    active_certificates = Certificate.objects.filter(is_active=True).count()
    revoked_certificates = total_certificates - active_certificates
    certificate_segments, certificate_circumference = _donut_segments([
        active_certificates, revoked_certificates,
    ])

    # Skill coverage — top skills by number of courses mapped to them
    skill_rows = (
        CourseSkill.objects
        .values(label=F('skill__name'))
        .annotate(value=Count('course', distinct=True))
        .order_by('-value')
    )
    skill_bars, skill_max = _simple_bar_series(skill_rows, 'label', 'value')

    # Learning streak distribution — bucket current streak lengths
    streak_buckets = [
        {'label': '0 days', 'value': LearningStreak.objects.filter(current_streak=0).count()},
        {'label': '1-3 days', 'value': LearningStreak.objects.filter(current_streak__gte=1, current_streak__lte=3).count()},
        {'label': '4-7 days', 'value': LearningStreak.objects.filter(current_streak__gte=4, current_streak__lte=7).count()},
        {'label': '8-14 days', 'value': LearningStreak.objects.filter(current_streak__gte=8, current_streak__lte=14).count()},
        {'label': '15+ days', 'value': LearningStreak.objects.filter(current_streak__gte=15).count()},
    ]
    streak_bars, streak_max = _simple_bar_series(streak_buckets, 'label', 'value', max_bars=5)
    streak_has_activity = any(b['value'] for b in streak_buckets)

    context = {
        'total_users': total_users,
        'total_courses': total_courses,
        'role_counts': role_counts,
        'active_courses': active_courses,
        'inactive_courses': inactive_courses,
        'enrollment_points': enrollment_points,
        'enrollment_has_activity': raw_max > 0,
        'enrollment_line_points': line_points_str,
        'enrollment_area_points': area_points_str,
        'enrollment_chart_w': chart_w,
        'enrollment_chart_h': chart_h,
        'enrollment_baseline_y': baseline_y,
        'enrollment_gridlines': enrollment_gridlines,
        'enrollment_last_point': enrollment_last_point,
        'enrollment_plot_top': pad_t,
        'role_segments': role_segments,
        'role_circumference': round(role_circumference, 2),
        'total_certificates': total_certificates,
        'active_certificates': active_certificates,
        'revoked_certificates': revoked_certificates,
        'certificate_segments': certificate_segments,
        'certificate_circumference': round(certificate_circumference, 2),
        'skill_bars': skill_bars,
        'skill_has_data': bool(skill_bars),
        'streak_bars': streak_bars,
        'streak_has_activity': streak_has_activity,
    }
    return render(request, 'adminpanel/dashboard.html', context)

@admin_required
def user_list(request):
    users = User.objects.select_related('profile').all()
    return render(request, 'adminpanel/users.html', {
        'users': users,
        'admin_create_form': AdminCreateForm(),
    })


@admin_required
def create_admin_user(request):
    if request.method == 'POST':
        form = AdminCreateForm(request.POST)
        if form.is_valid():
            new_admin = form.save()
            messages.success(request, f"{new_admin.username} has been added as an admin.")
            return redirect('admin_user_list')

        users = User.objects.select_related('profile').all()
        return render(request, 'adminpanel/users.html', {
            'users': users,
            'admin_create_form': form,
        })

    return redirect('admin_user_list')


@admin_required
def user_profile_detail(request, user_id):
    user = get_object_or_404(User.objects.select_related('profile'), id=user_id)
    return render(request, 'adminpanel/user_detail.html', {'profile_user': user})


@admin_required
def update_user_profile_detail(request, user_id):
    user = get_object_or_404(User.objects.select_related('profile'), id=user_id)

    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        user.email = request.POST.get('email', '').strip()
        user.save()

        profile = user.profile
        profile.phone_number = request.POST.get('phone_number', '').strip()
        profile.institution_name = request.POST.get('institution_name', '').strip()
        profile.linkedin_url = request.POST.get('linkedin_url', '').strip()
        profile.github_url = request.POST.get('github_url', '').strip()
        profile.xp = request.POST.get('xp') or 0
        profile.level = request.POST.get('level') or 1
        profile.skills = request.POST.get('skills', '').strip()

        if request.FILES.get('profile_image'):
            profile.profile_image = request.FILES.get('profile_image')

        profile.save()
        messages.success(request, f"{user.username}'s profile has been updated.")

    return redirect('admin_user_profile_detail', user_id=user.id)


@admin_required
def delete_user_account(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('admin_user_profile_detail', user_id=user.id)

    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"{username}'s account has been deleted.")
        return redirect('admin_user_list')

    return redirect('admin_user_profile_detail', user_id=user.id)


@admin_required
def update_user_role(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # prevent admin from demoting self
    if user == request.user:
        messages.error(request, "You cannot change your own role.")
        return redirect('admin_user_list')

    new_role = request.POST.get('role')
    if new_role in ['STUDENT', 'INSTRUCTOR']:
        user.profile.role = new_role
        user.profile.save()
        messages.success(request, f"{user.username}'s role updated to {new_role}")

    return redirect('admin_user_list')


@admin_required
def pending_approvals(request):
    users = User.objects.filter(
        profile__approval_status__in=['PENDING', 'REJECTED']
    ).select_related('profile').order_by('profile__approval_status', '-date_joined')
    return render(request, 'adminpanel/pending_approvals.html', {'users': users})


@admin_required
def approve_user(request, user_id):
    user = get_object_or_404(User.objects.select_related('profile'), id=user_id)

    user.is_active = True
    user.save()

    profile = user.profile
    profile.approval_status = 'APPROVED'
    profile.save()

    notify(user, "Your registration has been approved. You can now log in.", link=reverse('login'))
    messages.success(request, f"{user.username}'s registration has been approved.")

    return redirect(request.META.get('HTTP_REFERER') or 'admin_user_list')


@admin_required
def reject_user(request, user_id):
    user = get_object_or_404(User.objects.select_related('profile'), id=user_id)

    if user == request.user:
        messages.error(request, "You cannot reject your own account.")
        return redirect(request.META.get('HTTP_REFERER') or 'admin_user_list')

    user.is_active = False
    user.save()

    profile = user.profile
    profile.approval_status = 'REJECTED'
    profile.save()

    messages.success(request, f"{user.username}'s registration has been rejected.")

    return redirect(request.META.get('HTTP_REFERER') or 'admin_user_list')


@admin_required
def toggle_user_status(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # prevent admin from deactivating self
    if user == request.user:
        messages.error(request, "You cannot deactivate your own account.")
        return redirect('admin_user_list')

    user.is_active = not user.is_active
    user.save()

    if user.is_active and user.profile.approval_status != 'APPROVED':
        user.profile.approval_status = 'APPROVED'
        user.profile.save()

    status = "activated" if user.is_active else "deactivated"
    messages.success(request, f"{user.username} has been {status}.")

    return redirect('admin_user_list')

@admin_required
def course_list(request):
    courses = Course.objects.select_related('instructor').all()
    instructors = User.objects.filter(profile__role='INSTRUCTOR', is_active=True)
    return render(
        request,
        'adminpanel/courses.html',
        {
            'courses': courses,
            'instructors': instructors
        }
    )


@admin_required
def create_course(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        instructor_id = request.POST.get('instructor')

        instructor = User.objects.get(id=instructor_id) if instructor_id else None

        course = Course.objects.create(
            title=title,
            description=description,
            instructor=instructor,
            image=request.FILES.get('image')
        )

        notify_all_students(
            f"New course added: {course.title}",
            link=reverse('student_course_list')
        )

        return redirect('admin_course_list')


@admin_required
def toggle_course_status(request, course_id):
    course = Course.objects.get(id=course_id)
    course.is_active = not course.is_active
    course.save()
    return redirect('admin_course_list')

@admin_required
def lesson_list(request, course_id):
    course = Course.objects.get(id=course_id)
    lessons = course.lessons.all()
    return render(
        request,
        'adminpanel/lessons.html',
        {
            'course': course,
            'lessons': lessons
        }
    )


@admin_required
def create_lesson(request, course_id):
    course = Course.objects.get(id=course_id)

    if request.method == 'POST':
        Lesson.objects.create(
            course=course,
            title=request.POST.get('title'),
            content=request.POST.get('content'),
            resource_link=request.POST.get('resource_link'),
            order=request.POST.get('order', 1),
            image=request.FILES.get('image')
        )
    return redirect('admin_lesson_list', course_id=course.id)


@admin_required
def toggle_lesson_status(request, lesson_id):
    lesson = Lesson.objects.get(id=lesson_id)
    lesson.is_active = not lesson.is_active
    lesson.save()
    return redirect('admin_lesson_list', course_id=lesson.course.id)

@admin_required
def quiz_list(request, course_id):
    course = Course.objects.get(id=course_id)
    quizzes = course.quizzes.all()
    return render(
        request,
        'adminpanel/quizzes.html',
        {
            'course': course,
            'quizzes': quizzes
        }
    )


@admin_required
def create_quiz(request, course_id):
    course = Course.objects.get(id=course_id)

    if request.method == 'POST':
        Quiz.objects.create(
            course=course,
            title=request.POST.get('title'),
            difficulty=request.POST.get('difficulty'),
            time_limit=request.POST.get('time_limit'),
            total_marks=request.POST.get('total_marks'),
            pass_mark=request.POST.get('pass_mark')
        )
    return redirect('admin_quiz_list', course_id=course.id)


@admin_required
def toggle_quiz_status(request, quiz_id):
    quiz = Quiz.objects.get(id=quiz_id)
    quiz.is_active = not quiz.is_active
    quiz.save()
    return redirect('admin_quiz_list', course_id=quiz.course.id)

@admin_required
def question_list(request, quiz_id):
    quiz = Quiz.objects.get(id=quiz_id)
    questions = quiz.questions.all()
    return render(
        request,
        'adminpanel/questions.html',
        {
            'quiz': quiz,
            'questions': questions
        }
    )


@admin_required
def create_question(request, quiz_id):
    quiz = Quiz.objects.get(id=quiz_id)

    if request.method == 'POST':
        Question.objects.create(
            quiz=quiz,
            text=request.POST.get('text'),
            option_a=request.POST.get('option_a'),
            option_b=request.POST.get('option_b'),
            option_c=request.POST.get('option_c'),
            option_d=request.POST.get('option_d'),
            correct_option=request.POST.get('correct_option'),
        )
    return redirect('admin_question_list', quiz_id=quiz.id)

@admin_required
def progress_overview(request):
    progress_records = StudentProgress.objects.select_related(
        'student', 'course'
    ).all()

    return render(
        request,
        'adminpanel/progress.html',
        {
            'progress_records': progress_records
        }
    )
    
@admin_required
def achievement_list(request):
    achievements = Achievement.objects.all()
    return render(
        request,
        'adminpanel/achievements.html',
        {'achievements': achievements}
    )


@admin_required
def create_achievement(request):
    if request.method == 'POST':
        achievement = Achievement.objects.create(
            title=request.POST.get('title'),
            description=request.POST.get('description'),
            xp_required=request.POST.get('xp_required')
        )
        notify_all_students(f"New achievement added: {achievement.title}")
    return redirect('admin_achievement_list')


@admin_required
def toggle_achievement_status(request, achievement_id):
    achievement = Achievement.objects.get(id=achievement_id)
    achievement.is_active = not achievement.is_active
    achievement.save()
    return redirect('admin_achievement_list')

@admin_required
def streak_overview(request):
    streaks = LearningStreak.objects.select_related(
        'student'
    ).filter(student__profile__role='STUDENT')
    rewards = StreakReward.objects.all()

    return render(
        request,
        'adminpanel/streaks.html',
        {
            'streaks': streaks,
            'rewards': rewards
        }
    )


@admin_required
def create_streak_reward(request):
    if request.method == 'POST':
        reward = StreakReward.objects.create(
            streak_days=request.POST.get('streak_days'),
            bonus_xp=request.POST.get('bonus_xp')
        )
        notify_all_students(
            f"New streak reward added: {reward.streak_days}-day streak → {reward.bonus_xp} XP"
        )
    return redirect('admin_streaks')

@admin_required
def skill_list(request):
    skills = Skill.objects.all()
    courses = Course.objects.all()
    mappings = CourseSkill.objects.select_related('course', 'skill')

    return render(
        request,
        'adminpanel/skills.html',
        {
            'skills': skills,
            'courses': courses,
            'mappings': mappings
        }
    )


@admin_required
def create_skill(request):
    if request.method == 'POST':
        skill = Skill.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description')
        )
        notify_all_students(f"New skill added: {skill.name}")
    return redirect('admin_skill_list')


@admin_required
def map_course_skill(request):
    if request.method == 'POST':
        mapping = CourseSkill.objects.create(
            course_id=request.POST.get('course'),
            skill_id=request.POST.get('skill'),
            weight=request.POST.get('weight', 1)
        )
        notify_course_students(
            mapping.course,
            f"New skill added to {mapping.course.title}: {mapping.skill.name}"
        )
    return redirect('admin_skill_list')

@admin_required
def time_tracking_overview(request):
    sessions = LearningSession.objects.select_related(
        'student', 'course'
    ).order_by('-start_time')[:100]

    activities = ActivityLog.objects.select_related(
        'student', 'course'
    ).order_by('-timestamp')[:100]

    return render(
        request,
        'adminpanel/time_tracking.html',
        {
            'sessions': sessions,
            'activities': activities
        }
    )

from courses.models import Certificate

@admin_required
def certificate_list(request):
    certificates = (
        Certificate.objects
        .select_related('student', 'course')
        .order_by('-issued_at')
    )

    courses = Course.objects.all()

    course_id = request.GET.get('course')
    if course_id:
        certificates = certificates.filter(course_id=course_id)

    return render(
        request,
        'adminpanel/certificates.html',
        {
            'certificates': certificates,
            'courses': courses,
            'selected_course': course_id
        }
    )

@admin_required
def revoke_certificate(request, certificate_id):
    certificate = get_object_or_404(Certificate, id=certificate_id)
    certificate.is_active = False
    certificate.save()
    return redirect('admin_certificate_list')


@admin_required
def reissue_certificate(request, certificate_id):
    certificate = get_object_or_404(Certificate, id=certificate_id)
    certificate.is_active = True
    certificate.save()
    return redirect('admin_certificate_list')


@admin_required
def user_activity_log(request):
    rows = []
    if os.path.exists(LOG_FILE):
        from openpyxl import load_workbook
        workbook = load_workbook(LOG_FILE, read_only=True)
        sheet = workbook['User Activity']
        rows = [
            row for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and row[0] and row[6] == 'Register'
        ]
        rows.reverse()

    return render(request, 'adminpanel/user_activity_log.html', {
        'rows': rows,
        'log_exists': os.path.exists(LOG_FILE),
        'pending_count': pending_count(),
    })


@admin_required
def download_user_activity_log(request):
    if not os.path.exists(LOG_FILE):
        messages.error(request, "No activity has been logged yet.")
        return redirect('admin_user_activity_log')

    return FileResponse(
        open(LOG_FILE, 'rb'),
        as_attachment=True,
        filename='user_activity.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@admin_required
def live_user_activity_log(request):
    """Return the current contents of the workbook as JSON, read fresh from disk each call,
    so the frontend can show a live spreadsheet preview without downloading the file."""
    headers = []
    data_rows = []

    if os.path.exists(LOG_FILE):
        from openpyxl import load_workbook

        def cell_value(value):
            if value is None:
                return ''
            if hasattr(value, 'isoformat'):
                return value.isoformat()
            return value

        workbook = load_workbook(LOG_FILE, read_only=True, data_only=True)
        sheet = workbook['User Activity']
        all_rows = list(sheet.iter_rows(values_only=True))
        if all_rows:
            headers = [cell_value(h) or '' for h in all_rows[0]]
            data_rows = [
                [cell_value(c) for c in row]
                for row in all_rows[1:]
                if row and any(c is not None for c in row)
            ]
        workbook.close()

    return JsonResponse({
        'headers': headers,
        'rows': data_rows,
        'log_exists': os.path.exists(LOG_FILE),
    })


@admin_required
def backfill_user_activity_log(request):
    if request.method == 'POST':
        added = backfill_existing_users()
        if added:
            messages.success(request, f"Added {added} existing user(s) to the activity log.")
        else:
            messages.info(request, "Every user is already in the activity log.")

    return redirect('admin_user_activity_log')


@admin_required
def sync_user_activity_log(request):
    if request.method == 'POST':
        synced = sync_pending_entries()
        if synced:
            messages.success(request, f"Synced {synced} pending entr{'y' if synced == 1 else 'ies'} into the activity log.")
        else:
            messages.info(request, "Nothing to sync — the workbook was saved successfully or there's nothing pending.")

    return redirect('admin_user_activity_log')
